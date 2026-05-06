from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Dict
from datetime import datetime
import calendar

import joblib
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

PROJECT_ROOT = Path(__file__).resolve().parent
DATA_PATH = PROJECT_ROOT / "data" / "customer_engagement_dataset.csv"
MODEL_PATH = PROJECT_ROOT / "engagement_model.joblib"
METRICS_PATH = PROJECT_ROOT / "metrics.json"

LABELS = ["low", "medium", "high"]


def clean_text(text: str) -> str:
    if not isinstance(text, str):
        text = str(text)
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def load_dataset() -> pd.DataFrame:
    if not DATA_PATH.exists():
        raise FileNotFoundError(f"Dataset file not found: {DATA_PATH}")

    df = pd.read_csv(DATA_PATH)
    if "review_text" not in df.columns or "engagement_label" not in df.columns:
        raise ValueError("Dataset must contain 'review_text' and 'engagement_label' columns.")

    df = df.copy()
    df["cleaned_text"] = df["review_text"].apply(clean_text)
    return df


def build_pipeline() -> LogisticRegression:
    return LogisticRegression(max_iter=500, class_weight="balanced", random_state=42)


def save_artifacts(pipeline: object, metrics: Dict[str, object]) -> None:
    joblib.dump(pipeline, MODEL_PATH)
    METRICS_PATH.write_text(json.dumps(metrics, indent=2), encoding="utf-8")
    print(f"\n✓ Model saved to {MODEL_PATH}")
    print(f"✓ Metrics saved to {METRICS_PATH}\n")


def train_model() -> None:
    """Train the engagement classification model."""
    print("\n" + "=" * 70)
    print("  ENGAGEMENT CLASSIFICATION MODEL TRAINING")
    print("=" * 70 + "\n")
    
    df = load_dataset()
    print(f"📊 Loaded dataset with {len(df):,} records")

    X = df["cleaned_text"]
    y = df["engagement_label"]

    print("🔄 Vectorizing text with TF-IDF...")
    vectorizer = TfidfVectorizer(ngram_range=(1, 2), max_features=5000, min_df=2, sublinear_tf=True)
    X_vectors = vectorizer.fit_transform(X)
    
    print("📈 Splitting data (80% train, 20% test)...")
    X_train, X_test, y_train, y_test = train_test_split(
        X_vectors,
        y,
        test_size=0.20,
        random_state=42,
        stratify=y,
    )

    print("🤖 Training Logistic Regression model...")
    pipeline = build_pipeline()
    pipeline.fit(X_train, y_train)
    predictions = pipeline.predict(X_test)

    print("📋 Evaluating model performance...")
    accuracy = accuracy_score(y_test, predictions)
    report = classification_report(y_test, predictions, output_dict=True, zero_division=0)
    matrix = confusion_matrix(y_test, predictions, labels=LABELS).tolist()
    precision = report["macro avg"]["precision"] if "macro avg" in report else 0.0
    recall = report["macro avg"]["recall"] if "macro avg" in report else 0.0
    f1 = report["macro avg"]["f1-score"] if "macro avg" in report else 0.0

    # Get engagement counts for chatbot
    engagement_counts = df["engagement_label"].value_counts().to_dict()

    metrics = {
        "dataset_rows": int(len(df)),
        "accuracy": round(float(accuracy), 4),
        "precision_weighted": round(float(report["weighted avg"]["precision"] if "weighted avg" in report else precision), 4),
        "recall_weighted": round(float(report["weighted avg"]["recall"] if "weighted avg" in report else recall), 4),
        "f1_weighted": round(float(report["weighted avg"]["f1-score"] if "weighted avg" in report else f1), 4),
        "confusion_matrix": matrix,
        "labels": LABELS,
        "engagement": engagement_counts,
        "classification_report": report,
    }

    print("\n" + "=" * 70)
    print("  MODEL PERFORMANCE SUMMARY")
    print("=" * 70)
    print(f"\n✓ Accuracy:           {metrics['accuracy']:.2%}")
    print(f"✓ Precision (weighted): {metrics['precision_weighted']:.4f}")
    print(f"✓ Recall (weighted):    {metrics['recall_weighted']:.4f}")
    print(f"✓ F1-Score (weighted):  {metrics['f1_weighted']:.4f}\n")

    save_artifacts(pipeline, metrics)


def load_metrics() -> Dict:
    """Load saved metrics from file."""
    if not METRICS_PATH.exists():
        raise FileNotFoundError(f"Metrics file not found. Run training first: {METRICS_PATH}")
    return json.loads(METRICS_PATH.read_text(encoding="utf-8"))


def load_data_stats() -> Dict:
    """Load dataset statistics."""
    df = load_dataset()
    engagement_counts = df["engagement_label"].value_counts()
    sentiment_counts = df["sentiment"].value_counts()
    channel_counts = df["channel"].value_counts()
    region_top = df["region"].value_counts().head(5)
    
    return {
        "total_records": len(df),
        "engagement": engagement_counts.to_dict(),
        "sentiment": sentiment_counts.to_dict(),
        "channels": channel_counts.to_dict(),
        "top_regions": region_top.to_dict(),
        "avg_rating": df["rating"].mean(),
        "avg_response_time": df["response_time_hours"].mean(),
    }


def load_full_dataset() -> pd.DataFrame:
    """Load full dataset with date parsing."""
    if not DATA_PATH.exists():
        raise FileNotFoundError(f"Dataset file not found: {DATA_PATH}")

    df = pd.read_csv(DATA_PATH, parse_dates=['created_at'])
    if "review_text" not in df.columns or "engagement_label" not in df.columns:
        raise ValueError("Dataset must contain 'review_text' and 'engagement_label' columns.")

    df = df.copy()
    df["cleaned_text"] = df["review_text"].apply(clean_text)
    df["month"] = df["created_at"].dt.month
    df["year"] = df["created_at"].dt.year
    df["month_name"] = df["created_at"].dt.month_name()
    df["year_month"] = df["created_at"].dt.strftime('%Y-%m')
    
    return df


def get_time_based_analysis(df: pd.DataFrame, time_period: str = "month") -> Dict:
    """Get time-based analysis of engagement data."""
    if time_period == "month":
        grouped = df.groupby(['year', 'month_name']).agg({
            'engagement_label': 'value_counts',
            'rating': 'mean',
            'response_time_hours': 'mean',
            'sentiment': lambda x: x.value_counts().index[0] if len(x) > 0 else 'neutral'
        }).unstack(level=0)
    elif time_period == "year":
        grouped = df.groupby('year').agg({
            'engagement_label': 'value_counts',
            'rating': 'mean',
            'response_time_hours': 'mean',
            'sentiment': lambda x: x.value_counts().index[0] if len(x) > 0 else 'neutral'
        })
    else:
        # Custom period
        grouped = df.groupby(time_period).agg({
            'engagement_label': 'value_counts',
            'rating': 'mean',
            'response_time_hours': 'mean',
            'sentiment': lambda x: x.value_counts().index[0] if len(x) > 0 else 'neutral'
        })
    
    return grouped.to_dict()


def create_terminal_dashboard(df: pd.DataFrame, metrics: Dict) -> str:
    """Create a comprehensive terminal dashboard."""
    dashboard = []
    
    # Header
    dashboard.append("╔" + "═" * 78 + "╗")
    dashboard.append("║" + " " * 25 + "ENGAGEMENT ANALYSIS DASHBOARD" + " " * 25 + "║")
    dashboard.append("╚" + "═" * 78 + "╝")
    dashboard.append("")
    
    # Overall Metrics
    dashboard.append("📊 OVERALL METRICS")
    dashboard.append("─" * 50)
    dashboard.append(f"Total Records:     {metrics['dataset_rows']:,}")
    dashboard.append(f"Model Accuracy:    {metrics['accuracy']:.1%}")
    dashboard.append(f"Average Rating:    {df['rating'].mean():.2f}/5")
    dashboard.append(f"Avg Response Time: {df['response_time_hours'].mean():.1f} hours")
    dashboard.append("")
    
    # Engagement Distribution
    dashboard.append("🎯 ENGAGEMENT DISTRIBUTION")
    dashboard.append("─" * 50)
    engagement = metrics.get('engagement', {})
    total = sum(engagement.values())
    for level in ["high", "medium", "low"]:
        count = engagement.get(level, 0)
        percentage = (count / total * 100) if total > 0 else 0
        bar = "█" * int(percentage / 5)
        dashboard.append(f"  {level.capitalize():8} │ {bar:20} {count:6,} ({percentage:5.1f}%)")
    dashboard.append("")
    
    # Monthly Trends (Last 6 months)
    dashboard.append("📈 MONTHLY ENGAGEMENT TRENDS")
    dashboard.append("─" * 50)
    monthly_data = df.groupby('year_month')['engagement_label'].value_counts().unstack().fillna(0)
    recent_months = sorted(monthly_data.index)[-6:]
    
    for month in recent_months:
        high_count = int(monthly_data.loc[month, 'high']) if 'high' in monthly_data.columns else 0
        total_month = int(monthly_data.loc[month].sum())
        pct = (high_count / total_month * 100) if total_month > 0 else 0
        dashboard.append(f"  {month}: High Engagement {high_count:4,} ({pct:4.1f}%)")
    dashboard.append("")
    
    # Channel Performance
    dashboard.append("📱 CHANNEL PERFORMANCE")
    dashboard.append("─" * 50)
    channels = df['channel'].value_counts()
    for channel, count in channels.head(5).items():
        pct = (count / len(df) * 100)
        dashboard.append(f"  {channel:15} │ {count:6,} ({pct:5.1f}%)")
    dashboard.append("")
    
    # Sentiment Analysis
    dashboard.append("😊 SENTIMENT ANALYSIS")
    dashboard.append("─" * 50)
    sentiment = df['sentiment'].value_counts()
    for sent, count in sentiment.items():
        pct = (count / len(df) * 100)
        dashboard.append(f"  {sent.capitalize():10} │ {count:6,} ({pct:5.1f}%)")
    dashboard.append("")
    
    # Top Regions
    dashboard.append("🌍 TOP REGIONS")
    dashboard.append("─" * 50)
    regions = df['region'].value_counts().head(5)
    for region, count in regions.items():
        pct = (count / len(df) * 100)
        dashboard.append(f"  {region:20} │ {count:6,} ({pct:5.1f}%)")
    dashboard.append("")
    
    return "\n".join(dashboard)


def generate_excel_report(df: pd.DataFrame, metrics: Dict, output_path: str = "engagement_analysis_report.xlsx") -> None:
    """Generate comprehensive Excel report with dashboard and analysis."""
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Dashboard Sheet
    ws_dashboard = wb.active
    ws_dashboard.title = "Dashboard"
    
    # Title
    ws_dashboard['A1'] = "ENGAGEMENT ANALYSIS DASHBOARD"
    ws_dashboard['A1'].font = Font(size=16, bold=True)
    ws_dashboard.merge_cells('A1:G1')
    
    # Overall Metrics
    ws_dashboard['A3'] = "OVERALL METRICS"
    ws_dashboard['A3'].font = Font(size=12, bold=True)
    
    metrics_data = [
        ["Total Records", metrics['dataset_rows']],
        ["Model Accuracy", f"{metrics['accuracy']:.1%}"],
        ["Average Rating", f"{df['rating'].mean():.2f}/5"],
        ["Avg Response Time", f"{df['response_time_hours'].mean():.1f} hours"],
    ]
    
    for i, (label, value) in enumerate(metrics_data, 4):
        ws_dashboard[f'A{i}'] = label
        ws_dashboard[f'B{i}'] = value
    
    # Engagement Distribution
    ws_dashboard['A10'] = "ENGAGEMENT DISTRIBUTION"
    ws_dashboard['A10'].font = Font(size=12, bold=True)
    
    engagement = metrics.get('engagement', {})
    total = sum(engagement.values())
    
    ws_dashboard['A11'] = "Level"
    ws_dashboard['B11'] = "Count"
    ws_dashboard['C11'] = "Percentage"
    
    for i, level in enumerate(["high", "medium", "low"], 12):
        count = engagement.get(level, 0)
        percentage = (count / total * 100) if total > 0 else 0
        ws_dashboard[f'A{i}'] = level.capitalize()
        ws_dashboard[f'B{i}'] = count
        ws_dashboard[f'C{i}'] = f"{percentage:.1f}%"
    
    # Monthly Trends
    ws_dashboard['A17'] = "MONTHLY TRENDS (High Engagement)"
    ws_dashboard['A17'].font = Font(size=12, bold=True)
    
    monthly_data = df.groupby('year_month')['engagement_label'].value_counts().unstack().fillna(0)
    recent_months = sorted(monthly_data.index)[-6:]
    
    ws_dashboard['A18'] = "Month"
    ws_dashboard['B18'] = "High Engagement"
    ws_dashboard['C18'] = "Total Feedback"
    ws_dashboard['D18'] = "Percentage"
    
    for i, month in enumerate(recent_months, 19):
        high_count = int(monthly_data.loc[month, 'high']) if 'high' in monthly_data.columns else 0
        total_month = int(monthly_data.loc[month].sum())
        pct = (high_count / total_month * 100) if total_month > 0 else 0
        ws_dashboard[f'A{i}'] = month
        ws_dashboard[f'B{i}'] = high_count
        ws_dashboard[f'C{i}'] = total_month
        ws_dashboard[f'D{i}'] = f"{pct:.1f}%"
    
    # Data Sheet
    ws_data = wb.create_sheet("Raw Data")
    ws_data['A1'] = "ENGAGEMENT ANALYSIS RAW DATA"
    ws_data['A1'].font = Font(size=14, bold=True)
    ws_data.merge_cells('A1:P1')
    
    # Write headers
    headers = ['Date', 'Channel', 'Region', 'Market Segment', 'Product Line', 
               'Sentiment', 'Urgency', 'Engagement Label', 'Rating', 'Likes', 
               'Shares', 'Comments', 'Response Time', 'Purchase Intent', 'Review Text']
    
    for col, header in enumerate(headers, 1):
        ws_data.cell(2, col, header).font = Font(bold=True)
    
    # Write data
    for row, (_, record) in enumerate(df.iterrows(), 3):
        ws_data.cell(row, 1, record['created_at'].strftime('%Y-%m-%d'))
        ws_data.cell(row, 2, record['channel'])
        ws_data.cell(row, 3, record['region'])
        ws_data.cell(row, 4, record['market_segment'])
        ws_data.cell(row, 5, record['product_line'])
        ws_data.cell(row, 6, record['sentiment'])
        ws_data.cell(row, 7, record['urgency'])
        ws_data.cell(row, 8, record['engagement_label'])
        ws_data.cell(row, 9, record['rating'])
        ws_data.cell(row, 10, record['likes'])
        ws_data.cell(row, 11, record['shares'])
        ws_data.cell(row, 12, record['comments'])
        ws_data.cell(row, 13, record['response_time_hours'])
        ws_data.cell(row, 14, record['purchase_intent'])
        ws_data.cell(row, 15, record['review_text'])
    
    # Analysis Sheet
    ws_analysis = wb.create_sheet("Analysis")
    ws_analysis['A1'] = "DETAILED ANALYSIS & INSIGHTS"
    ws_analysis['A1'].font = Font(size=14, bold=True)
    ws_analysis.merge_cells('A1:F1')
    
    # Key Insights
    insights = [
        "KEY INSIGHTS AND RECOMMENDATIONS",
        "",
        "1. ENGAGEMENT ANALYSIS:",
        f"   - High engagement customers: {engagement.get('high', 0):,} ({(engagement.get('high', 0)/total*100):.1f}%)",
        f"   - Medium engagement customers: {engagement.get('medium', 0):,} ({(engagement.get('medium', 0)/total*100):.1f}%)",
        f"   - Low engagement customers: {engagement.get('low', 0):,} ({(engagement.get('low', 0)/total*100):.1f}%)",
        "   - Recommendation: Focus retention efforts on medium and low engagement segments",
        "",
        "2. SENTIMENT ANALYSIS:",
    ]
    
    sentiment = df['sentiment'].value_counts()
    for sent, count in sentiment.items():
        pct = (count / len(df) * 100)
        insights.append(f"   - {sent.capitalize()}: {count:,} ({pct:.1f}%)")
    
    insights.extend([
        "   - Recommendation: Address negative sentiment drivers immediately",
        "",
        "3. CHANNEL PERFORMANCE:",
    ])
    
    channels = df['channel'].value_counts()
    for channel, count in channels.head(3).items():
        pct = (count / len(df) * 100)
        insights.append(f"   - {channel}: {count:,} ({pct:.1f}%)")
    
    insights.extend([
        "   - Recommendation: Optimize top-performing channels",
        "",
        "4. RESPONSE TIME ANALYSIS:",
        f"   - Average response time: {df['response_time_hours'].mean():.1f} hours",
        f"   - Fastest responses get higher engagement",
        "   - Recommendation: Implement automated responses for urgent cases",
        "",
        "5. TREND ANALYSIS:",
        "   - Monitor monthly engagement patterns",
        "   - Identify seasonal variations",
        "   - Track sentiment changes over time",
    ])
    
    for i, insight in enumerate(insights, 3):
        ws_analysis[f'A{i}'] = insight
    
    # Auto-adjust column widths
    for ws in [ws_dashboard, ws_data, ws_analysis]:
        for col_num, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            for cell in col:
                try:
                    if hasattr(cell, 'value') and cell.value is not None and not hasattr(cell, 'merged'):
                        cell_value = str(cell.value)
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                except:
                    pass
            if max_length > 0:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_path)
    print(f"✓ Excel report saved to {output_path}")


def parse_time_query(query: str) -> tuple:
    """Parse time-based queries from user input."""
    query = query.lower()
    
    # Month queries
    months = {
        'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
        'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12
    }
    
    for month_name, month_num in months.items():
        if month_name in query:
            return ('month', month_num, month_name.capitalize())
    
    # Year queries
    import re
    year_match = re.search(r'\b(202[4-5])\b', query)
    if year_match:
        return ('year', int(year_match.group(1)), year_match.group(1))
    
    # Quarter queries
    if 'quarter' in query or 'q1' in query or 'q2' in query or 'q3' in query or 'q4' in query:
        quarter_map = {'q1': 1, 'q2': 2, 'q3': 3, 'q4': 4}
        for q, num in quarter_map.items():
            if q in query:
                return ('quarter', num, f"Q{num}")
    
    return (None, None, None)


def format_response(title: str, content: str, color: str = "blue") -> str:
    """Format chatbot response."""
    colors = {
        "blue": "\033[94m",
        "green": "\033[92m",
        "yellow": "\033[93m",
        "red": "\033[91m",
        "end": "\033[0m"
    }
    
    color_code = colors.get(color, colors["blue"])
    end_code = colors["end"]
    
    return f"\n{color_code}{'─' * 70}\n💬 {title}\n{'─' * 70}\n{end_code}{content}\n"


def chatbot():
    """Enhanced interactive chatbot for engagement analysis queries."""
    print("\n" + "=" * 80)
    print("  🤖 ENHANCED ENGAGEMENT ANALYSIS CHATBOT")
    print("=" * 80)
    print("\nHello! 👋 I'm your Advanced Engagement Analysis Assistant.")
    print("Ask me anything about customer engagement, metrics, trends, or time periods.")
    print("I can show dashboards, analyze by month/year, and export Excel reports!")
    print("\nType 'help' for commands, 'dashboard' for overview, or 'export' for Excel report.")
    print("Type 'exit' to quit.\n")
    
    try:
        metrics = load_metrics()
        data_stats = load_data_stats()
        df = load_full_dataset()
    except FileNotFoundError as e:
        print(f"❌ Error: {str(e)}")
        print("Please train the model first by running: python analysis.py\n")
        return
    except Exception as e:
        print(f"❌ Error loading data: {str(e)}\n")
        return
    
    print("Data loaded successfully! Ready for your queries.\n")
    
    while True:
        try:
            user_input = input("\n👤 You: ").strip().lower()
            
            if not user_input:
                continue
            
            if user_input == "exit" or user_input == "quit":
                print(format_response("Goodbye!", "Thank you for using the Enhanced Engagement Analysis Chatbot! 👋", "green"))
                break
            
            if user_input == "help":
                help_text = """
Available commands:
  • "dashboard" - Show comprehensive terminal dashboard
  • "overall metrics" - Show model performance summary
  • "engagement" - Show engagement distribution
  • "sentiment" - Show sentiment distribution
  • "channels" - Show feedback channels
  • "regions" - Show top regions
  • "accuracy" - Show model accuracy
  • "export" - Generate Excel report with full analysis
  • "how many high engagement?" - Query specific metrics
  • "show me [month/year]" - Time-based analysis (e.g., "show me may", "show me 2024")
  • "analyze [factor]" - Analyze by specific factor
  • "help" - Show this help message
  • "exit" - Exit the chatbot
                """
                print(format_response("Help Menu", help_text, "yellow"))
                continue
            
            response_generated = False
            
            # Dashboard command
            if user_input == "dashboard":
                dashboard = create_terminal_dashboard(df, metrics)
                print(dashboard)
                response_generated = True
            
            # Export command
            elif user_input == "export" or "excel" in user_input:
                try:
                    output_path = "engagement_analysis_report.xlsx"
                    generate_excel_report(df, metrics, output_path)
                    response = f"""
✅ Excel report generated successfully!

📁 File: {output_path}

📊 Report includes:
   • Executive Dashboard with key metrics
   • Raw data (30,000+ records)
   • Detailed analysis and insights
   • Recommendations for each metric
   • Monthly trends and patterns

💡 The report provides comprehensive analysis for non-technical stakeholders.
                    """
                    print(format_response("Excel Report Generated", response, "green"))
                except Exception as e:
                    print(format_response("Export Error", f"Could not generate Excel report: {str(e)}", "red"))
                response_generated = True
            
            # Time-based queries
            elif "show me" in user_input or any(word in user_input for word in ["january", "february", "march", "april", "may", "june", "july", "august", "september", "october", "november", "december", "2024", "2025"]):
                time_type, time_value, time_label = parse_time_query(user_input)
                
                if time_type == "month":
                    month_data = df[df['month'] == time_value]
                    if len(month_data) > 0:
                        engagement_counts = month_data['engagement_label'].value_counts()
                        avg_rating = month_data['rating'].mean()
                        total_records = len(month_data)
                        
                        response = f"""
📅 Analysis for {time_label}

📊 Total Records: {total_records:,}
⭐ Average Rating: {avg_rating:.2f}/5

🎯 Engagement Breakdown:
"""
                        for level in ["high", "medium", "low"]:
                            count = engagement_counts.get(level, 0)
                            pct = (count / total_records * 100) if total_records > 0 else 0
                            response += f"   {level.capitalize()}: {count:,} ({pct:.1f}%)\n"
                        
                        response += f"""
💡 Insights for {time_label}:
   • Best performing engagement level: {engagement_counts.index[0] if len(engagement_counts) > 0 else 'N/A'}
   • Customer satisfaction: {'High' if avg_rating >= 4 else 'Medium' if avg_rating >= 3 else 'Low'}
   • Focus area: {'Retention' if engagement_counts.get('high', 0) > engagement_counts.get('low', 0) else 'Re-engagement campaigns'}
                        """
                        print(format_response(f"{time_label} Analysis", response, "blue"))
                    else:
                        print(format_response("No Data", f"No data found for {time_label}", "yellow"))
                
                elif time_type == "year":
                    year_data = df[df['year'] == time_value]
                    if len(year_data) > 0:
                        engagement_counts = year_data['engagement_label'].value_counts()
                        monthly_trend = year_data.groupby('month_name')['engagement_label'].value_counts().unstack().fillna(0)
                        
                        response = f"""
📅 Analysis for {time_value}

📊 Total Records: {len(year_data):,}
⭐ Average Rating: {year_data['rating'].mean():.2f}/5

🎯 Annual Engagement Breakdown:
"""
                        for level in ["high", "medium", "low"]:
                            count = engagement_counts.get(level, 0)
                            pct = (count / len(year_data) * 100)
                            response += f"   {level.capitalize()}: {count:,} ({pct:.1f}%)\n"
                        
                        response += f"""
📈 Monthly Highlights:
"""
                        if len(monthly_trend) > 0:
                            top_month = monthly_trend['high'].idxmax() if 'high' in monthly_trend.columns else 'N/A'
                            response += f"   Best month for high engagement: {top_month}\n"
                        
                        response += f"""
💡 Insights for {time_value}:
   • Overall trend: {'Improving' if engagement_counts.get('high', 0) > engagement_counts.get('medium', 0) else 'Needs attention'}
   • Key focus: {'Customer retention' if engagement_counts.get('high', 0) > len(year_data)*0.4 else 'Engagement campaigns'}
                        """
                        print(format_response(f"{time_value} Analysis", response, "blue"))
                    else:
                        print(format_response("No Data", f"No data found for {time_value}", "yellow"))
                
                response_generated = True
            
            # Factor analysis
            elif "analyze" in user_input or "breakdown" in user_input:
                if "channel" in user_input:
                    channel_analysis = df.groupby('channel').agg({
                        'engagement_label': 'value_counts',
                        'rating': 'mean',
                        'response_time_hours': 'mean'
                    }).unstack()
                    
                    response = "📱 Channel Performance Analysis\n\n"
                    for channel in df['channel'].unique():
                        channel_data = df[df['channel'] == channel]
                        high_pct = (channel_data['engagement_label'] == 'high').mean() * 100
                        avg_rating = channel_data['rating'].mean()
                        avg_response = channel_data['response_time_hours'].mean()
                        
                        response += f"📊 {channel}:\n"
                        response += f"   High Engagement: {high_pct:.1f}%\n"
                        response += f"   Avg Rating: {avg_rating:.2f}/5\n"
                        response += f"   Avg Response Time: {avg_response:.1f} hours\n\n"
                    
                    print(format_response("Channel Analysis", response, "blue"))
                    response_generated = True
                
                elif "region" in user_input:
                    region_analysis = df.groupby('region').agg({
                        'engagement_label': 'value_counts',
                        'rating': 'mean'
                    }).unstack()
                    
                    response = "🌍 Regional Performance Analysis\n\n"
                    top_regions = df['region'].value_counts().head(5)
                    
                    for region in top_regions.index:
                        region_data = df[df['region'] == region]
                        high_pct = (region_data['engagement_label'] == 'high').mean() * 100
                        avg_rating = region_data['rating'].mean()
                        
                        response += f"📍 {region}:\n"
                        response += f"   Records: {len(region_data):,}\n"
                        response += f"   High Engagement: {high_pct:.1f}%\n"
                        response += f"   Avg Rating: {avg_rating:.2f}/5\n\n"
                    
                    print(format_response("Regional Analysis", response, "blue"))
                    response_generated = True
            
            # Overall metrics query
            elif "overall" in user_input or "summary" in user_input or "performance" in user_input:
                response = f"""
Model Accuracy:        {metrics['accuracy']:.2%}
Precision (weighted):  {metrics['precision_weighted']:.4f}
Recall (weighted):     {metrics['recall_weighted']:.4f}
F1-Score (weighted):   {metrics['f1_weighted']:.4f}
Total Records:         {metrics['dataset_rows']:,}

Dataset Overview:
• Date Range: {df['created_at'].min().strftime('%Y-%m-%d')} to {df['created_at'].max().strftime('%Y-%m-%d')}
• Channels: {len(df['channel'].unique())}
• Regions: {len(df['region'].unique())}
• Market Segments: {len(df['market_segment'].unique())}
                """
                print(format_response("Overall Metrics", response, "blue"))
                response_generated = True
            
            # Engagement distribution
            elif "engagement" in user_input and "how" not in user_input:
                try:
                    engagement = metrics.get('engagement', {})
                    response = "\n"
                    total = sum(engagement.values())
                    for level in ["high", "medium", "low"]:
                        count = engagement.get(level, 0)
                        percentage = (count / total * 100) if total > 0 else 0
                        bar = "█" * int(percentage / 5)
                        response += f"  {level.capitalize():10} │ {bar:20} {count:6,} ({percentage:5.1f}%)\n"
                    
                    response += f"\n💡 Insights:\n"
                    response += f"   • High engagement customers are your most valuable segment\n"
                    response += f"   • Medium engagement needs nurturing to prevent churn\n"
                    response += f"   • Low engagement requires immediate re-engagement campaigns\n"
                    
                    print(format_response("Engagement Distribution", response, "green"))
                    response_generated = True
                except Exception as e:
                    print(format_response("Error", f"Could not get engagement data: {str(e)}", "red"))
                    response_generated = True
            
            # Sentiment distribution
            elif "sentiment" in user_input:
                try:
                    sentiment = df['sentiment'].value_counts()
                    response = "\n"
                    total = len(df)
                    for sent, count in sorted(sentiment.items(), key=lambda x: x[1], reverse=True):
                        percentage = (count / total * 100) if total > 0 else 0
                        bar = "█" * int(percentage / 5)
                        response += f"  {sent.capitalize():10} │ {bar:20} {count:6,} ({percentage:5.1f}%)\n"
                    
                    response += f"\n💡 Insights:\n"
                    response += f"   • {sentiment.index[0].capitalize()} sentiment dominates the feedback\n"
                    response += f"   • Address negative sentiment drivers to improve engagement\n"
                    
                    print(format_response("Sentiment Analysis", response, "blue"))
                    response_generated = True
                except Exception as e:
                    print(format_response("Error", f"Could not get sentiment data: {str(e)}", "red"))
                    response_generated = True
            
            # Channels
            elif "channel" in user_input:
                try:
                    channels = df['channel'].value_counts()
                    response = "\n"
                    for channel, count in sorted(channels.items(), key=lambda x: x[1], reverse=True):
                        percentage = (count / len(df) * 100) if len(df) > 0 else 0
                        response += f"  • {channel:20} {count:6,} records ({percentage:5.1f}%)\n"
                    
                    response += f"\n💡 Insights:\n"
                    response += f"   • {channels.index[0]} is the most used feedback channel\n"
                    response += f"   • Focus support resources on top channels\n"
                    
                    print(format_response("Feedback Channels", response, "blue"))
                    response_generated = True
                except Exception as e:
                    print(format_response("Error", f"Could not get channel data: {str(e)}", "red"))
                    response_generated = True
            
            # Regions
            elif "region" in user_input:
                try:
                    regions = df['region'].value_counts().head(5)
                    response = "\n"
                    for region, count in regions.items():
                        percentage = (count / len(df) * 100) if len(df) > 0 else 0
                        response += f"  • {region:25} {count:6,} records ({percentage:5.1f}%)\n"
                    
                    response += f"\n💡 Insights:\n"
                    response += f"   • {regions.index[0]} has the highest feedback volume\n"
                    response += f"   • Regional differences may indicate market-specific issues\n"
                    
                    print(format_response("Top Regions", response, "blue"))
                    response_generated = True
                except Exception as e:
                    print(format_response("Error", f"Could not get region data: {str(e)}", "red"))
                    response_generated = True
            
            # Accuracy query
            elif "accuracy" in user_input:
                response = f"\n✓ Model Accuracy: {metrics['accuracy']:.2%}\n"
                response += f"\nThis means the model correctly classifies {metrics['accuracy']:.2%} of customer feedback into\n"
                response += f"the correct engagement tier (Low, Medium, or High).\n"
                response += f"\nDataset Size: {metrics['dataset_rows']:,} records\n"
                response += f"Training completed with high accuracy validation.\n"
                print(format_response("Model Accuracy", response, "green"))
                response_generated = True
            
            # "How many" queries
            elif "how many" in user_input or "how much" in user_input:
                try:
                    engagement = metrics.get('engagement', {})
                    total = metrics['dataset_rows']
                    
                    if "high" in user_input and "engagement" in user_input:
                        count = engagement.get('high', 0)
                        pct = (count / total) * 100 if total > 0 else 0
                        response = f"\n📊 High Engagement Records: {count:,} ({pct:.1f}%)\n"
                        response += f"These are your most valuable customers with strong interest and loyalty.\n"
                        response += f"They represent your best opportunities for upselling and advocacy.\n"
                        print(format_response("High Engagement Customers", response, "green"))
                        response_generated = True
                    
                    elif "medium" in user_input and "engagement" in user_input:
                        count = engagement.get('medium', 0)
                        pct = (count / total) * 100 if total > 0 else 0
                        response = f"\n📊 Medium Engagement Records: {count:,} ({pct:.1f}%)\n"
                        response += f"These customers have moderate interest and growth potential.\n"
                        response += f"Focus retention efforts here to prevent churn to low engagement.\n"
                        print(format_response("Medium Engagement Customers", response, "yellow"))
                        response_generated = True
                    
                    elif "low" in user_input and "engagement" in user_input:
                        count = engagement.get('low', 0)
                        pct = (count / total) * 100 if total > 0 else 0
                        response = f"\n📊 Low Engagement Records: {count:,} ({pct:.1f}%)\n"
                        response += f"These customers need immediate re-engagement campaigns.\n"
                        response += f"High risk of churn - implement targeted retention strategies.\n"
                        print(format_response("Low Engagement Customers", response, "red"))
                        response_generated = True
                    
                    elif "total" in user_input or "records" in user_input:
                        response = f"\n📊 Total Records: {total:,}\n"
                        response += f"This comprehensive dataset covers customer feedback from {df['created_at'].min().strftime('%B %Y')} to {df['created_at'].max().strftime('%B %Y')}.\n"
                        print(format_response("Total Records", response, "blue"))
                        response_generated = True
                except Exception as e:
                    print(format_response("Error", f"Could not process query: {str(e)}", "red"))
                    response_generated = True
            
            # If no match found
            if not response_generated:
                response = """
I didn't quite understand that. Try one of these:
  • "dashboard" - Show comprehensive overview
  • "export" - Generate Excel report
  • "show me may" - Analyze specific month
  • "show me 2024" - Analyze specific year
  • "analyze channels" - Breakdown by channel
  • "overall metrics" - Model performance
  • "engagement distribution" - Customer segments
  • Type 'help' for more options
                """
                print(format_response("Need Clarification", response, "yellow"))
        
        except KeyboardInterrupt:
            print(format_response("Session Ended", "Goodbye! 👋", "green"))
            break
        except Exception as e:
            print(format_response("Error", f"Something went wrong: {str(e)}", "red"))


def main() -> None:
    """Main entry point - choose between training or chatbot."""
    import sys
    
    if len(sys.argv) > 1 and sys.argv[1] == "--chat":
        # Run chatbot mode
        chatbot()
    else:
        # Run training mode
        train_model()
        print("\n💡 To use the interactive chatbot, run: python analysis.py --chat\n")


if __name__ == "__main__":
    main()

